#!/usr/bin/env python3
"""
Đọc testcase từ file JSON và ghi vào Excel
Giải quyết vấn đề vượt quá giới hạn độ dài dòng lệnh
"""

import argparse
import json
import os
import sys
import zipfile
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
except ImportError:
    print("❌ Thiếu thư viện openpyxl, vui lòng chạy: pip install openpyxl")
    sys.exit(1)


def get_config(templates_dir):
    """Tải cấu hình template"""
    config_path = templates_dir / "template-config.json"
    if config_path.exists():
        with open(config_path, 'r', encoding='utf-8') as f:
            return json.load(f)
    return None


def read_xlsx_template_columns(template_path):
    """Đọc định nghĩa cột từ template Excel"""
    columns = []
    
    # Sử dụng openpyxl đọc trực tiếp
    from openpyxl import load_workbook
    wb = load_workbook(template_path)
    ws = wb.active
    
    for cell in ws[1]:
        if cell.value:
            columns.append({
                "index": cell.column - 1,
                "name": cell.value
            })
    
    return columns


def merge_columns(xlsx_columns, config_columns):
    """Gộp tiêu đề cột Excel và định nghĩa cột cấu hình"""
    name_to_key = {col['name']: col['key'] for col in config_columns}
    name_to_required = {col['name']: col.get('required', False) for col in config_columns}
    
    result = []
    for col in xlsx_columns:
        key = name_to_key.get(col['name'], col['name'].lower().replace(' ', '_'))
        required = name_to_required.get(col['name'], False)
        result.append({
            "index": col['index'],
            "name": col['name'],
            "key": key,
            "required": required
        })
    return result


def format_steps(value):
    """Chuyển đổi các bước phân tách thành định dạng nhiều dòng"""
    if not isinstance(value, str):
        return value
    # Xử lý thẻ xuống dòng HTML <br> → \n
    value = value.replace('<br>', '\n').replace('<br/>', '\n').replace('<BR>', '\n')

    # Nếu đã là nhiều dòng, dọn dẹp trực tiếp
    if '\n' in value:
        lines = [line.strip() for line in value.split('\n') if line.strip()]
        return '\n'.join(lines)

    # Nếu chứa dấu chấm phẩy (tiếng Trung hoặc tiếng Anh), thay thế bằng xuống dòng
    if '；' in value or ';' in value:
        value = value.replace('；', '\n').replace(';', '\n')
        lines = [line.strip() for line in value.split('\n') if line.strip()]
        return '\n'.join(lines)

    # Nếu ở định dạng số thứ tự (ví dụ: "1. xxx 2. xxx 3. xxx"), phân tách theo số thứ tự
    import re
    # Khớp định dạng "Số. " hoặc "Số、"
    pattern = r'(\d+[.、])\s*'
    # Chèn ký tự xuống dòng trước mỗi "Số. "
    value = re.sub(pattern, r'\n\1', value).strip()
    if value.startswith('\n'):
        value = value[1:]

    return value


def write_excel(output_path, columns, testcases, title="Testcase"):
    """Ghi file Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = title
    
    header_font = Font(bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    wrap_alignment = Alignment(wrap_text=True, vertical='top')
    
    # Ghi tiêu đề
    for col in columns:
        cell = ws.cell(row=1, column=col['index'] + 1, value=col['name'])
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Ghi dữ liệu
    for row_idx, tc in enumerate(testcases, 2):
        for col in columns:
            key = col['key']
            value = tc.get(key, "")
            # Xử lý kiểu danh sách/từ điển
            if isinstance(value, (list, dict)):
                value = json.dumps(value, ensure_ascii=False)
            # Định dạng các bước
            value = format_steps(value)
            cell = ws.cell(row=row_idx, column=col['index'] + 1, value=str(value) if value else "")
            cell.border = thin_border
            cell.alignment = wrap_alignment
    
    # Thiết lập độ rộng cột
    for col in columns:
        col_letter = chr(64 + col['index'] + 1)
        ws.column_dimensions[col_letter].width = 20
    
    ws.freeze_panes = 'A2'
    wb.save(output_path)
    return len(testcases)


def main():
    parser = argparse.ArgumentParser(description='Đọc testcase từ file JSON và ghi vào Excel')
    parser.add_argument('--data', '-d', required=True, help='Đường dẫn file JSON (chứa mảng testcases)')
    parser.add_argument('--template', '-t', default='Bản_mẫu_testcase.xlsx', help='Tên template Excel')
    parser.add_argument('--output', '-o', required=True, help='Đường dẫn xuất file Excel')
    parser.add_argument('--title', default='Testcase', help='Tiêu đề sheet')
    parser.add_argument('--template-key', default='default', help='Key template trong template-config.json')
    
    args = parser.parse_args()
    
    # Kiểm tra file JSON
    if not os.path.exists(args.data):
        print(f"❌ File JSON không tồn tại: {args.data}")
        sys.exit(1)
    
    # Xử lý đường dẫn
    script_dir = Path(os.path.dirname(os.path.abspath(__file__)))
    templates_dir = script_dir.parent / 'assets'
    
    # 1. Tải dữ liệu JSON
    print(f"📄 Đang đọc file dữ liệu: {args.data}")
    with open(args.data, 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    # Hỗ trợ mảng trực tiếp hoặc định dạng {testcases: [...]}
    testcases = data if isinstance(data, list) else data.get('testcases', [])
    print(f"   Số lượng testcase: {len(testcases)}")
    
    # 2. Tải cấu hình
    config = get_config(templates_dir)
    if not config:
        print("❌ Không tìm thấy template-config.json")
        sys.exit(1)
    
    config_columns = config.get('templates', {}).get(args.template_key, {}).get('columns', [])
    
    # 3. Đọc template Excel
    template_path = templates_dir / args.template
    if not template_path.exists():
        print(f"❌ File template không tồn tại: {template_path}")
        sys.exit(1)
    
    xlsx_columns = read_xlsx_template_columns(str(template_path))
    columns = merge_columns(xlsx_columns, config_columns)
    print(f"   Số lượng cột template: {len(columns)}")
    
    # 4. Ghi file Excel
    output_path = os.path.abspath(args.output)
    count = write_excel(output_path, columns, testcases, args.title)
    
    print(f"\n✅ Excel đã được tạo: {output_path}")
    print(f"   Số lượng testcase: {count}")


if __name__ == '__main__':
    main()
