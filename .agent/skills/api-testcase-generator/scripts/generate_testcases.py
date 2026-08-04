#!/usr/bin/env python3
"""
API Testcase Generator
Tạo ca kiểm thử (testcase) ở nhiều định dạng khác nhau dựa trên kết quả phân tích JSON của API.
Hỗ trợ: pytest, postman, jmeter, excel
"""

import json
import argparse
import os
import re
from pathlib import Path
from urllib.parse import urlparse
from jinja2 import Template


# ========== Định nghĩa template ==========

PYTEST_TEMPLATE = '''"""
Auto-generated API tests
Generated from: {{ info.title or 'Swagger API' }}
Base URL: {{ base_url }}
"""
import pytest
import requests

BASE_URL = "{{ base_url }}"
{% if auth_type == 'bearer' %}
TOKEN = "your_token_here"  # TODO: configure auth token
{% endif %}


class Test{{ resource_title }}:
    """{{ resource_desc }}"""

{% for case in cases %}
    def {{ case.func_name }}(self):
        """{{ case.description }}"""
        {% if case.setup_code %}
{{ case.setup_code }}
        {% endif %}
        {% if case.path_params %}
        path = "{{ case.path }}".format({{ case.path_params }})
        {% else %}
        path = "{{ case.path }}"
        {% endif %}
        url = f"{BASE_URL}{path}"
        {% if case.content_type == 'application/x-www-form-urlencoded' %}
        response = requests.{{ case.method.lower() }}(url, data=payload, headers=headers)
        {% elif case.content_type and 'multipart' in case.content_type %}
        response = requests.{{ case.method.lower() }}(url, files=payload, headers=headers)
        {% elif case.method in ['POST', 'PUT', 'PATCH'] %}
        response = requests.{{ case.method.lower() }}(url, json=payload, headers=headers)
        {% elif case.method == 'GET' %}
        response = requests.{{ case.method.lower() }}(url, params=params, headers=headers)
        {% else %}
        response = requests.{{ case.method.lower() }}(url, headers=headers)
        {% endif %}
        {% if case.assert_status %}
        assert response.status_code == {{ case.assert_status }}
        {% endif %}
        {% if case.assertions %}
        data = response.json()
        {% for assertion in case.assertions %}
{{ assertion }}
        {% endfor %}
        {% endif %}
        {% if case.extract_var %}
        # Extract variable for subsequent requests
        {{ case.extract_var }} = data.get("{{ case.extract_field }}")
        {% endif %}
{% endfor %}
'''

POSTMAN_TEMPLATE = '''{
  "info": {
    "name": "{{ info.title or 'API Tests' }}",
    "description": "Auto-generated from Swagger",
    "schema": "https://schema.getpostman.com/json/collection/v2.1.0/collection.json"
  },
  "item": [
{% for scenario in scenarios %}
    {
      "name": "{{ scenario.name }}",
      "item": [
{% for api in scenario.apis %}
        {
          "name": "{{ api.summary or api.path }}",
          "request": {
            "method": "{{ api.method }}",
            "header": [
{% if auth_type == 'bearer' %}
              {
                "key": "Authorization",
                "value": "Bearer {{token}}",
                "type": "text"
              }
{% endif %}
            ],
            "url": {
              "raw": "{{ base_url }}{{ api.path }}",
              "host": ["{{ base_url }}"],
              "path": [{% for seg in api.path_segments %}"{{ seg }}"{% if not loop.last %}, {% endif %}{% endfor %}]
            }
          },
          "event": [
            {
              "listen": "test",
              "script": {
                "exec": [
{% if api.assert_status %}
                  "pm.test('Status code is {{ api.assert_status }}', function () {",
                  "  pm.response.to.have.status({{ api.assert_status }});",
                  "});",
{% endif %}
{% for test in api.tests %}
                  "{{ test }}",
{% endfor %}
                  ""
                ],
                "type": "text/javascript"
              }
            }
          ]
        }{% if not loop.last %},{% endif %}
{% endfor %}
      ]
    }{% if not loop.last %},{% endif %}
{% endfor %}
  ]
}
'''

JMETER_TEMPLATE = '''<?xml version="1.0" encoding="UTF-8"?>
<jmeterTestPlan version="1.2" properties="5.0" jmeter="5.6">
  <hashTree>
    <TestPlan guiclass="TestPlanGui" testclass="TestPlan" testname="{{ info.title or 'API Tests' }}" enabled="true">
      <stringProp name="TestPlan.comments">Auto-generated from Swagger</stringProp>
    </TestPlan>
    <hashTree>
      <ThreadGroup guiclass="ThreadGroupGui" testclass="ThreadGroup" testname="API Scenario" enabled="true">
        <elementProp name="ThreadGroup.arguments" elementType="Arguments">
          <collectionProp name="Arguments.arguments"/>
        </elementProp>
      </ThreadGroup>
      <hashTree>
{% for api in apis %}
        <HTTPSamplerProxy guiclass="HttpTestSampleGui" testclass="HTTPSamplerProxy" testname="{{ api.summary or api.path }}" enabled="true">
          <elementProp name="HTTPsampler.Arguments" elementType="Arguments">
            <collectionProp name="Arguments.arguments"/>
          </elementProp>
          <stringProp name="HTTPSampler.domain">{{ domain }}</stringProp>
          <stringProp name="HTTPSampler.port">{{ port }}</stringProp>
          <stringProp name="HTTPSampler.protocol">{{ scheme }}</stringProp>
          <stringProp name="HTTPSampler.path">{{ api.path }}</stringProp>
          <stringProp name="HTTPSampler.method">{{ api.method }}</stringProp>
        </HTTPSamplerProxy>
        <hashTree>
          <HeaderManager guiclass="HeaderPanel" testclass="HeaderManager" testname="HTTP Headers" enabled="true">
            <collectionProp name="HeaderManager.headers">
{% if auth_type == 'bearer' %}
              <elementProp name="" elementType="Header">
                <stringProp name="Header.name">Authorization</stringProp>
                <stringProp name="Header.value">Bearer ${TOKEN}</stringProp>
              </elementProp>
{% endif %}
            </collectionProp>
          </HeaderManager>
          <hashTree/>
{% if api.assert_status %}
          <ResponseAssertion guiclass="AssertionGui" testclass="ResponseAssertion" testname="Status Assertion" enabled="true">
            <collectionProp name="Asserion.test_strings">
              <stringProp name="">{{ api.assert_status }}</stringProp>
            </collectionProp>
            <stringProp name="Assertion.test_field">Assertion.response_code</stringProp>
          </ResponseAssertion>
          <hashTree/>
{% endif %}
        </hashTree>
{% endfor %}
      </hashTree>
    </hashTree>
  </hashTree>
</jmeterTestPlan>
'''


def load_parsed(input_path):
    with open(input_path, 'r', encoding='utf-8') as f:
        return json.load(f)


def build_base_url(data):
    scheme = data.get('schemes', ['https'])[0]
    host = data.get('host', '') or 'localhost'
    base_path = data.get('basePath', '')
    port = ''
    if ':' in host:
        parts = host.split(':')
        host = parts[0]
        port = f':{parts[1]}'
    return f"{scheme}://{host}{port}{base_path}"


def generate_pytest(data, output_dir, scenarios=None):
    """Sinh mã nguồn Pytest"""
    base_url = build_base_url(data)
    apis = data['apis']
    scenarios = scenarios or data.get('scenarios', [])

    # Gom nhóm theo tag hoặc resource
    groups = {}
    for api in apis:
        key = api['tags'][0] if api['tags'] else api['resource'] or 'default'
        if key not in groups:
            groups[key] = []
        groups[key].append(api)

    os.makedirs(output_dir, exist_ok=True)

    for group, g_apis in groups.items():
        cases = []
        for api in g_apis:
            case = build_case(api, 'pytest')
            cases.append(case)

        resource_title = ''.join(w.capitalize() for w in re.split(r'[^a-zA-Z0-9]', group) if w)
        template = Template(PYTEST_TEMPLATE)
        content = template.render(
            info=data.get('info', {}),
            base_url=base_url,
            auth_type='bearer',  # TODO: tự động nhận diện
            resource_title=resource_title,
            resource_desc=group,
            cases=cases
        )

        filename = f"test_{group.lower().replace(' ', '_')}.py"
        filepath = os.path.join(output_dir, filename)
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(content)
        print(f"Generated: {filepath}")


def generate_postman(data, output_path, scenarios=None):
    """Sinh Postman Collection"""
    base_url = build_base_url(data)
    scenarios = scenarios or data.get('scenarios', [])

    # Nếu không có kịch bản, đưa tất cả API vào một kịch bản mặc định
    if not scenarios:
        scenarios = [{
            'name': 'All APIs',
            'apis': data['apis']
        }]
    else:
        # Tìm kiếm định nghĩa API đầy đủ dựa trên API ID của kịch bản
        api_map = {a['operation_id'] or a['path']: a for a in data['apis']}
        for s in scenarios:
            s['apis'] = [api_map.get(aid, {'path': aid, 'method': 'GET', 'summary': aid}) for aid in s['apis']]

    for s in scenarios:
        for api in s['apis']:
            api['path_segments'] = [p for p in api['path'].split('/') if p]
            api['tests'] = build_case(api, 'postman').get('tests', [])
            api['assert_status'] = infer_status_code(api)

    template = Template(POSTMAN_TEMPLATE)
    content = template.render(
        info=data.get('info', {}),
        base_url=base_url,
        auth_type='bearer',
        token='{{token}}',
        scenarios=scenarios
    )

    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(content)
    print(f"Generated: {output_path}")


def generate_jmeter(data, output_path):
    """Sinh cấu hình JMeter JMX"""
    base_url = build_base_url(data)
    parsed = urlparse(base_url)

    apis = data['apis']
    for api in apis:
        api['assert_status'] = infer_status_code(api)

    template = Template(JMETER_TEMPLATE)
    content = template.render(
        info=data.get('info', {}),
        domain=parsed.hostname or 'localhost',
        port=parsed.port or ('443' if parsed.scheme == 'https' else '80'),
        scheme=parsed.scheme,
        auth_type='bearer',
        apis=apis
    )

    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(content)
    print(f"Generated: {output_path}")


def generate_excel(data, output_path):
    """Sinh testcase Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        print("Error: openpyxl not installed. Run: pip install openpyxl")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "API_Testcase"

    headers = ["Mã testcase", "Tên kịch bản", "Phân hệ/Module", "Đường dẫn API", "Phương thức",
               "Tiêu đề testcase", "Loại testcase", "Độ ưu tiên", "Điều kiện tiên quyết", "Các bước thực hiện", "Kết quả mong muốn", "Ghi chú"]
    ws.append(headers)

    # Định dạng
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    scenarios = data.get('scenarios', [])
    if not scenarios:
        scenarios = [{'name': 'Kiểm thử API đơn lẻ', 'apis': [a['operation_id'] or a['path'] for a in data['apis']]}]

    api_map = {a['operation_id'] or a['path']: a for a in data['apis']}
    row_idx = 2

    for s_idx, s in enumerate(scenarios, 1):
        for a_idx, aid in enumerate(s['apis'], 1):
            api = api_map.get(aid, {})
            case_id = f"TC_API_{s_idx:03d}_{a_idx:03d}"
            module = (api.get('tags', ['Module mặc định'])[0]) if api else 'Module mặc định'
            path = api.get('path', aid) if api else aid
            method = api.get('method', 'GET') if api else 'GET'
            title = api.get('summary', path) if api else path
            case_type = "Kiểm thử kịch bản" if len(s['apis']) > 1 else "Kiểm thử API đơn lẻ"
            priority = "P0" if s.get('type') == 'crud' and a_idx <= 2 else "P1"
            precondition = build_precondition(api, s, a_idx)
            steps = build_steps(api, s, a_idx)
            expected = build_expected(api, s, a_idx)
            remark = f"Kịch bản: {s.get('name', '')}"

            ws.append([case_id, s.get('name', ''), module, path, method,
                       title, case_type, priority, precondition, steps, expected, remark])
            row_idx += 1

    # Tự động căn chỉnh độ rộng cột
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 60)
        ws.column_dimensions[column].width = adjusted_width

    wb.save(output_path)
    print(f"Generated: {output_path}")


# ========== Hàm bổ trợ ==========

def infer_status_code(api):
    """Dự đoán mã trạng thái mong đợi"""
    method = api['method']
    if method == 'POST':
        return 201 if '201' in api.get('responses', {}) else 200
    elif method == 'DELETE':
        return 204 if '204' in api.get('responses', {}) else 200
    elif method == 'GET':
        return 200
    else:
        return 200


def build_case(api, fmt):
    """Xây dựng dữ liệu ca kiểm thử đơn lẻ"""
    method = api['method']
    status = infer_status_code(api)
    func_name = f"test_{api['resource']}_{method.lower()}"
    # Đảm bảo tên hàm là duy nhất
    path_parts = [p for p in api['path'].split('/') if p and not p.startswith('{')]
    if api['resource']:
        path_suffix = '_'.join(path_parts[-2:]) if len(path_parts) >= 2 else ''
        if path_suffix and path_suffix != api['resource'].replace('-', '_'):
            func_name = f"test_{path_suffix}_{method.lower()}"
    description = api['summary'] or f"{method} {api['path']}"

    result = {
        'func_name': func_name,
        'description': description,
        'path': api['path'],
        'method': method,
        'assert_status': status,
        'setup_code': '',
        'assertions': [],
        'tests': [],
        'extract_var': None,
        'extract_field': None,
        'content_type': None,
        'path_params': '',
    }

    # content_type từ request_body
    req_body = api.get('request_body')
    if req_body and req_body.get('content_type'):
        result['content_type'] = req_body['content_type']

    # path_params: format dạng dict nếu đường dẫn chứa {param}
    path_params = api.get('path_params', [])
    if path_params:
        result['path_params'] = ', '.join([f"{p}='<{p}>'" for p in path_params])

    # Sinh các khẳng định (assertion) dựa trên loại phương thức
    if method == 'POST':
        result['assertions'].append("        assert 'id' in data or 'data' in data")
        result['extract_var'] = 'created_id'
        result['extract_field'] = 'id'
    elif method == 'GET':
        result['assertions'].append("        assert data is not None")
    elif method in ('PUT', 'PATCH'):
        result['assertions'].append("        assert data is not None")
    elif method == 'DELETE':
        result['assertions'] = []  # Thường không có body phản hồi

    # Postman tests
    if fmt == 'postman':
        if method == 'POST':
            result['tests'].append("pm.test('Response has id', function () { var jsonData = pm.response.json(); pm.expect(jsonData).to.have.property('id'); });")
        elif method == 'GET':
            result['tests'].append("pm.test('Response is valid', function () { var jsonData = pm.response.json(); pm.expect(jsonData).to.not.be.undefined; });")

    # Xử lý tham số
    params = api.get('parameters', [])
    has_query = any(p['in'] == 'query' for p in params)
    has_header = any(p['in'] == 'header' for p in params)
    has_path = any(p['in'] == 'path' for p in params)
    has_request_body = api.get('request_body') is not None

    parts = []
    if has_query:
        parts.append("        params = {}  # TODO: điền tham số query")
    else:
        parts.append("        params = {}")
    if has_request_body:
        parts.append("        payload = {}  # TODO: điền request body")
    else:
        parts.append("        payload = {}")
    if has_header:
        parts.append("        headers = {}  # TODO: điền headers")
    else:
        parts.append("        headers = {}")
    result['setup_code'] = "\n".join(parts)

    return result


def build_precondition(api, scenario, step_idx):
    """Xây dựng mô tả điều kiện tiên quyết"""
    if step_idx == 1:
        if scenario.get('type') == 'auth':
            return "Đã hoàn thành đăng nhập, lấy Token hợp lệ"
        return "Dịch vụ API đang chạy bình thường, dữ liệu kiểm thử đã được chuẩn bị"
    return f"Bước trước {step_idx-1} thực hiện thành công, dữ liệu đã được sinh ra"


def build_steps(api, scenario, step_idx):
    """Xây dựng mô tả bước thực hiện"""
    if not api:
        return f"Bước {step_idx}: Gọi API"
    method = api.get('method', 'GET')
    path = api.get('path', '')
    params = api.get('parameters', [])
    param_desc = ', '.join([f"{p['name']}({p['in']})" for p in params[:3]]) if params else 'không'
    return f"Bước {step_idx}: Gửi yêu cầu {method} đến {path}, tham số: {param_desc}"


def build_expected(api, scenario, step_idx):
    """Xây dựng mô tả kết quả mong muốn"""
    if not api:
        return "Kỳ vọng: API phản hồi chính xác"
    status = infer_status_code(api)
    method = api.get('method', 'GET')
    expected = f"Mã trạng thái: {status}"

    if method == 'POST':
        expected += "; Phản hồi chứa dữ liệu được tạo và trường id"
    elif method == 'GET':
        if '_assert_404' in str(api):
            expected += "; Tài nguyên không tồn tại"
        else:
            expected += "; Phản hồi chứa dữ liệu chính xác"
    elif method in ('PUT', 'PATCH'):
        expected += "; Dữ liệu đã được cập nhật"
    elif method == 'DELETE':
        expected += "; Tài nguyên đã được xóa"

    return expected


def main():
    parser = argparse.ArgumentParser(description='Sinh API testcase')
    parser.add_argument('--input', '-i', required=True, help='Đường dẫn file JSON API đã phân tích')
    parser.add_argument('--format', '-f', required=True,
                        choices=['pytest', 'postman', 'jmeter', 'excel'],
                        help='Định dạng đầu ra')
    parser.add_argument('--output', '-o', required=True, help='Thư mục hoặc tệp đầu ra')
    parser.add_argument('--scenarios', '-s', help='Đường dẫn file JSON kịch bản tùy chỉnh (không bắt buộc)')
    args = parser.parse_args()

    data = load_parsed(args.input)
    scenarios = None
    if args.scenarios:
        with open(args.scenarios, 'r', encoding='utf-8') as f:
            scenarios = json.load(f)

    if args.format == 'pytest':
        generate_pytest(data, args.output, scenarios)
    elif args.format == 'postman':
        generate_postman(data, args.output, scenarios)
    elif args.format == 'jmeter':
        generate_jmeter(data, args.output)
    elif args.format == 'excel':
        generate_excel(data, args.output)


if __name__ == '__main__':
    main()
